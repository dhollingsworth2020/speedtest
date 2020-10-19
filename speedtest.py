import os
from subprocess import run
from  openpyxl import load_workbook, workbook
from datetime import datetime

raw_output = run(['cmd', '/C', os.path.join(os.path.dirname(__file__), 'speedtest.exe')], shell=True, capture_output=True)
parse_output = raw_output.stdout
split_parse_output = parse_output.split(b'\r\n')

dest_filename = os.path.join(os.path.dirname(__file__), 'speedtest.xlsx')
wb = load_workbook(filename=dest_filename)
last_row = wb.active.max_row
new_row = last_row + 1

extra = [b'Latency', b'Download', b'Upload', b'Packet Loss',b'Result URL']
spo = [split_parse_output[line].lstrip().rstrip().split(b':') for line in range(3,10)]
spo1 = []
for item in spo:
    for title in extra:
        if title in item[0]:
                if title == b'Result URL':
                    spo1.append(item[1]+item[2])
                elif title == b'Packet Loss':
                    if b'Not available' in item[1]:
                        spo1.append(b'Not available')    
                    else:
                        item_split = item[1].lstrip().rstrip().split(b' ')
                        spo1.append(item_split[0])
                        spo1.append(item_split[1])
                else:
                    item_split = item[1].lstrip().rstrip().split(b' ')
                    spo1.append(item_split[0])
                    spo1.append(item_split[1])
                break
    else:
        spo1.append(item[1])
            
spo1.insert(0,datetime.now())
spo1 = iter(spo1) 

for column in wb.active.iter_rows(min_row=new_row, max_col=11, max_row=new_row):
    for cell in column:
        try:
            cell.value = next(spo1)
        except:pass

wb.save(filename=dest_filename)
